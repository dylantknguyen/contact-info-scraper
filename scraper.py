import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re

# environment variables
LINK='https://apps.senate.virginia.gov/Senator/index.php'
FILE_NAME='info.xlsx'
SHEET_NAME='VA Senate'
NAME_SELECTOR='h3'
DEBUG=False

# get page
page = requests.get(LINK)
content = page.text

if DEBUG:
    print(content)

# scraper init
scraper  = BeautifulSoup(page.content, 'html.parser')
body = scraper.body

# workbook initialization
workbook = load_workbook(FILE_NAME)
worksheet = workbook.create_sheet(SHEET_NAME)

for link in scraper.find_all('a'):
    row_num = 1
    link = "https:" + link
    page = requests.get(link)

    name = scraper.find(NAME_SELECTOR).text
    phone_numbers = list(re.findall(r"((?:\d{3}|\(\d{3}\)){1}(?:\s|-|\.)?\d{3}(?:\s|-|\.)\d{4})", content))
    emails = list(re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,3}", content))

    if DEBUG:
        print(name)
        print(phone_numbers)
        print(emails)

    worksheet.cell(row=row_num, column=1, value=name)
    worksheet.cell(row=row_num, column=2, value=phone_numbers[0])
    worksheet.cell(row=row_num, column=3, value=emails[0])
    row_num+=1

workbook.save(FILE_NAME)
